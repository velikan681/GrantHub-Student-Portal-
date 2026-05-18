import json
from datetime import timedelta
from io import BytesIO

from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.views import LoginView, PasswordResetView
from django.core.exceptions import PermissionDenied
from django.db import IntegrityError
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .forms import (
    ApplicationForm,
    ApplicationStatusForm,
    CategoryForm,
    GrantForm,
    LoginForm,
    ProfileForm,
    StudentRegistrationForm,
)
from .models import Application, Category, Grant, Notification, SavedGrant, User
from .services import create_deadline_notifications, notify_matching_new_grant, recommended_grants_for


def is_manager(user):
    return user.is_authenticated and (user.is_staff or user.role in [User.Role.ADMIN, User.Role.ORGANIZATION])


def is_admin(user):
    return user.is_authenticated and (user.is_staff or user.role == User.Role.ADMIN)


def management_scope(user):
    grants = Grant.objects.select_related('category', 'created_by').annotate(applications_total=Count('applications'))
    applications = Application.objects.select_related('user', 'grant', 'grant__category', 'grant__created_by')
    if user.role == User.Role.ORGANIZATION and not user.is_staff:
        grants = grants.filter(created_by=user)
        applications = applications.filter(grant__created_by=user)
    return grants, applications


def filtered_grants(request):
    grants = Grant.objects.select_related('category', 'created_by').all()
    query = request.GET.get('q', '').strip()
    country = request.GET.get('country', '').strip()
    category = request.GET.get('category', '').strip()
    education_level = request.GET.get('education_level', '').strip()
    opportunity_type = request.GET.get('opportunity_type', '').strip()
    deadline = request.GET.get('deadline', '').strip()

    if query:
        grants = grants.filter(
            Q(title__icontains=query)
            | Q(description__icontains=query)
            | Q(organization__icontains=query)
            | Q(country__icontains=query)
            | Q(category__name__icontains=query)
        )
    if country:
        grants = grants.filter(country__icontains=country)
    if category:
        grants = grants.filter(category_id=category)
    if education_level:
        grants = grants.filter(education_level__icontains=education_level)
    if opportunity_type:
        grants = grants.filter(opportunity_type=opportunity_type)
    if deadline:
        grants = grants.filter(deadline__lte=deadline)

    sort = request.GET.get('sort', 'deadline')
    if sort == 'country':
        grants = grants.order_by('country', 'deadline')
    elif sort == 'popular':
        grants = grants.annotate(applications_total=Count('applications')).order_by('-applications_total', 'deadline')
    elif sort == 'new':
        grants = grants.order_by('-created_at')
    else:
        grants = grants.order_by('deadline', 'title')

    return grants


def home(request):
    actual = Grant.objects.select_related('category').filter(deadline__gte=timezone.localdate())[:6]
    soon = [grant for grant in actual if grant.is_deadline_soon][:4]
    categories = Category.objects.all()[:8]
    return render(request, 'portal/home.html', {'actual_grants': actual, 'soon_grants': soon, 'categories': categories})


def grants_list(request):
    grants = filtered_grants(request)
    context = {
        'grants': grants,
        'categories': Category.objects.all(),
        'types': Grant.OpportunityType.choices,
        'countries': Grant.objects.values_list('country', flat=True).distinct().order_by('country'),
    }
    return render(request, 'portal/grants_list.html', context)


def grant_detail(request, pk):
    grant = get_object_or_404(Grant.objects.select_related('category', 'created_by'), pk=pk)
    saved = request.user.is_authenticated and SavedGrant.objects.filter(user=request.user, grant=grant).exists()
    applied = request.user.is_authenticated and Application.objects.filter(user=request.user, grant=grant).exists()
    return render(request, 'portal/grant_detail.html', {'grant': grant, 'saved': saved, 'applied': applied})


def register(request):
    if request.method == 'POST':
        form = StudentRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, 'Регистрация завершена. Добро пожаловать в GrantHub!')
            return redirect('profile')
    else:
        form = StudentRegistrationForm()
    return render(request, 'portal/register.html', {'form': form})


class PortalLoginView(LoginView):
    template_name = 'portal/login.html'
    authentication_form = LoginForm


class PortalPasswordResetView(PasswordResetView):
    template_name = 'portal/password_reset.html'
    email_template_name = 'portal/password_reset_email.html'
    success_url = reverse_lazy('password_reset_done')


@login_required
def profile(request):
    create_deadline_notifications(request.user)
    if request.method == 'POST':
        form = ProfileForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Профиль обновлен.')
            return redirect('profile')
    else:
        form = ProfileForm(instance=request.user)
    return render(request, 'portal/profile.html', {
        'form': form,
        'applications': request.user.applications.select_related('grant')[:5],
        'saved_grants': SavedGrant.objects.filter(user=request.user).select_related('grant')[:5],
        'recommendations': recommended_grants_for(request.user),
        'notifications': request.user.notifications.all()[:6],
        'applications_count': request.user.applications.count(),
        'saved_count': SavedGrant.objects.filter(user=request.user).count(),
        'recommendations_count': recommended_grants_for(request.user).count(),
    })


@login_required
def save_grant(request, pk):
    grant = get_object_or_404(Grant, pk=pk)
    SavedGrant.objects.get_or_create(user=request.user, grant=grant)
    messages.success(request, 'Грант добавлен в сохраненные.')
    return redirect(request.POST.get('next') or 'grant_detail', pk=pk)


@login_required
def saved_grants(request):
    items = SavedGrant.objects.filter(user=request.user).select_related('grant', 'grant__category')
    return render(request, 'portal/saved_grants.html', {'items': items})


@login_required
def apply_grant(request, pk):
    grant = get_object_or_404(Grant, pk=pk)
    existing = Application.objects.filter(user=request.user, grant=grant).first()
    if existing:
        messages.info(request, 'Вы уже подали заявку на эту программу.')
        return redirect('my_applications')
    if request.method == 'POST':
        form = ApplicationForm(request.POST, request.FILES)
        if form.is_valid():
            application = form.save(commit=False)
            application.user = request.user
            application.grant = grant
            application.save()
            messages.success(request, 'Заявка отправлена и находится на рассмотрении.')
            return redirect('my_applications')
    else:
        form = ApplicationForm()
    return render(request, 'portal/apply.html', {'form': form, 'grant': grant})


@login_required
def my_applications(request):
    applications = Application.objects.filter(user=request.user).select_related('grant', 'grant__category')
    return render(request, 'portal/my_applications.html', {'applications': applications})


@login_required
def recommendations(request):
    return render(request, 'portal/recommendations.html', {'grants': recommended_grants_for(request.user)})


@login_required
def notifications(request):
    items = request.user.notifications.all()
    items.update(is_read=True)
    return render(request, 'portal/notifications.html', {'notifications': items})


@user_passes_test(is_manager)
def dashboard(request):
    grants, applications = management_scope(request.user)
    status = request.GET.get('status', '').strip()
    grant_id = request.GET.get('grant', '').strip()
    if status:
        applications = applications.filter(status=status)
    if grant_id:
        applications = applications.filter(grant_id=grant_id)

    base_applications = management_scope(request.user)[1]
    grants_count = grants.count()
    applications_count = base_applications.count()
    review_count = base_applications.filter(status=Application.Status.REVIEW).count()
    approved_count = base_applications.filter(status=Application.Status.APPROVED).count()
    rejected_count = base_applications.filter(status=Application.Status.REJECTED).count()
    soon_deadlines_count = grants.filter(deadline__gte=timezone.localdate(), deadline__lte=timezone.localdate() + timedelta(days=14)).count()
    unique_students_count = base_applications.values('user_id').distinct().count()
    popular_grants = grants.order_by('-applications_total', 'deadline')[:5]
    category_stats = grants.values('category__name').annotate(total=Count('id')).order_by('-total', 'category__name')[:8]
    country_stats = grants.values('country').annotate(total=Count('id')).order_by('-total', 'country')[:8]
    return render(request, 'portal/dashboard.html', {
        'grants': grants.order_by('deadline')[:20],
        'applications': applications.order_by('-submitted_at')[:25],
        'users_count': User.objects.count() if is_admin(request.user) else unique_students_count,
        'organizations_count': User.objects.filter(role=User.Role.ORGANIZATION).count() if is_admin(request.user) else 1,
        'students_count': User.objects.filter(role=User.Role.STUDENT).count() if is_admin(request.user) else unique_students_count,
        'grants_count': grants_count,
        'applications_count': applications_count,
        'review_count': review_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
        'soon_deadlines_count': soon_deadlines_count,
        'unique_students_count': unique_students_count,
        'popular_grants': popular_grants,
        'category_stats': category_stats,
        'country_stats': country_stats,
        'status_choices': Application.Status.choices,
        'selected_status': status,
        'selected_grant': grant_id,
        'categories': Category.objects.all(),
        'all_grants': grants.order_by('title'),
    })


def add_sheet(workbook, title, headers, rows):
    sheet = workbook.create_sheet(title)
    header_fill = PatternFill('solid', fgColor='EEF2FF')
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True, color='1F2937')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for row in rows:
        sheet.append(row)
    for column in sheet.columns:
        width = max(len(str(cell.value or '')) for cell in column) + 3
        sheet.column_dimensions[get_column_letter(column[0].column)].width = min(width, 42)
    sheet.freeze_panes = 'A2'
    return sheet


@user_passes_test(is_manager)
def export_excel(request):
    grants, applications = management_scope(request.user)
    workbook = Workbook()
    workbook.remove(workbook.active)

    add_sheet(workbook, 'Гранты', [
        'ID', 'Название', 'Организация', 'Страна', 'Категория', 'Тип',
        'Уровень образования', 'Дедлайн', 'Заявок', 'Автор', 'Создано',
    ], [
        [
            grant.id, grant.title, grant.organization, grant.country, grant.category.name,
            grant.get_opportunity_type_display(), grant.education_level,
            grant.deadline.strftime('%d.%m.%Y'), grant.applications_total,
            grant.created_by.full_name if grant.created_by else '-', grant.created_at.strftime('%d.%m.%Y %H:%M'),
        ]
        for grant in grants.order_by('deadline', 'title')
    ])

    add_sheet(workbook, 'Заявки', [
        'ID', 'Студент', 'Email', 'Университет', 'Факультет', 'Курс',
        'Грант', 'Организация', 'Страна', 'Категория', 'Статус', 'Дата подачи',
    ], [
        [
            app.id, app.user.full_name, app.user.email, app.user.university,
            app.user.faculty, app.user.course, app.grant.title, app.grant.organization,
            app.grant.country, app.grant.category.name, app.get_status_display(),
            app.submitted_at.strftime('%d.%m.%Y %H:%M'),
        ]
        for app in applications.order_by('-submitted_at')
    ])

    add_sheet(workbook, 'Статистика', ['Показатель', 'Значение'], [
        ['Грантов', grants.count()],
        ['Заявок всего', applications.count()],
        ['На рассмотрении', applications.filter(status=Application.Status.REVIEW).count()],
        ['Одобрено', applications.filter(status=Application.Status.APPROVED).count()],
        ['Отклонено', applications.filter(status=Application.Status.REJECTED).count()],
        ['Уникальных студентов', applications.values('user_id').distinct().count()],
        ['Дедлайн в ближайшие 14 дней', grants.filter(deadline__gte=timezone.localdate(), deadline__lte=timezone.localdate() + timedelta(days=14)).count()],
    ])

    if is_admin(request.user):
        add_sheet(workbook, 'Пользователи', [
            'ID', 'ФИО', 'Email', 'Роль', 'Университет', 'Факультет', 'Курс', 'Уровень', 'Интересы',
        ], [
            [
                user.id, user.full_name, user.email, user.get_role_display(), user.university,
                user.faculty, user.course, user.education_level, user.interests,
            ]
            for user in User.objects.order_by('role', 'full_name')
        ])

    stream = BytesIO()
    workbook.save(stream)
    filename = 'granthub-admin-report.xlsx' if is_admin(request.user) else 'granthub-organization-report.xlsx'
    response = HttpResponse(
        stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@user_passes_test(is_manager)
def grant_create(request):
    if request.method == 'POST':
        form = GrantForm(request.POST)
        if form.is_valid():
            grant = form.save(commit=False)
            grant.created_by = request.user
            grant.save()
            notify_matching_new_grant(grant)
            messages.success(request, 'Грантовая программа добавлена.')
            return redirect('dashboard')
    else:
        form = GrantForm()
    return render(request, 'portal/grant_form.html', {'form': form, 'title': 'Добавление гранта'})


@user_passes_test(is_manager)
def grant_update(request, pk):
    grant = get_object_or_404(Grant, pk=pk)
    if request.user.role == User.Role.ORGANIZATION and grant.created_by != request.user and not request.user.is_staff:
        raise PermissionDenied
    if request.method == 'POST':
        form = GrantForm(request.POST, instance=grant)
        if form.is_valid():
            form.save()
            messages.success(request, 'Грант обновлен.')
            return redirect('dashboard')
    else:
        form = GrantForm(instance=grant)
    return render(request, 'portal/grant_form.html', {'form': form, 'title': 'Редактирование гранта'})


@user_passes_test(is_admin)
def grant_delete(request, pk):
    grant = get_object_or_404(Grant, pk=pk)
    if request.method == 'POST':
        grant.delete()
        messages.success(request, 'Грант удален.')
    return redirect('dashboard')


@user_passes_test(is_manager)
def manage_applications(request):
    grants, applications = management_scope(request.user)
    status = request.GET.get('status', '').strip()
    grant_id = request.GET.get('grant', '').strip()
    if status:
        applications = applications.filter(status=status)
    if grant_id:
        applications = applications.filter(grant_id=grant_id)
    return render(request, 'portal/manage_applications.html', {
        'applications': applications.order_by('-submitted_at'),
        'status_form': ApplicationStatusForm(),
        'status_choices': Application.Status.choices,
        'selected_status': status,
        'selected_grant': grant_id,
        'all_grants': grants.order_by('title'),
    })


@user_passes_test(is_manager)
def update_application_status(request, pk):
    application = get_object_or_404(Application, pk=pk)
    if request.user.role == User.Role.ORGANIZATION and application.grant.created_by != request.user and not request.user.is_staff:
        raise PermissionDenied
    if request.method == 'POST':
        form = ApplicationStatusForm(request.POST, instance=application)
        if form.is_valid():
            form.save()
            Notification.objects.create(
                user=application.user,
                message=f'Статус заявки на "{application.grant.title}" изменен: {application.get_status_display()}.',
            )
            messages.success(request, 'Статус заявки обновлен.')
    return redirect('manage_applications')


@user_passes_test(is_admin)
def manage_categories(request):
    form = CategoryForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Направление добавлено.')
        return redirect('manage_categories')
    return render(request, 'portal/manage_categories.html', {'form': form, 'categories': Category.objects.all()})


@user_passes_test(is_admin)
def users_list(request):
    return render(request, 'portal/users_list.html', {'users': User.objects.all().order_by('role', 'full_name')})


def about(request):
    return render(request, 'portal/about.html')


def contacts(request):
    return render(request, 'portal/contacts.html')


def grant_to_dict(grant):
    return {
        'id': grant.id,
        'title': grant.title,
        'description': grant.description,
        'organization': grant.organization,
        'country': grant.country,
        'category': grant.category.name,
        'opportunity_type': grant.opportunity_type,
        'education_level': grant.education_level,
        'deadline': grant.deadline.isoformat(),
        'official_link': grant.official_link,
    }


@csrf_exempt
def api_grants(request, pk=None):
    if pk:
        grant = get_object_or_404(Grant, pk=pk)
        if request.method == 'GET':
            return JsonResponse(grant_to_dict(grant))
        if request.method in ['PUT', 'POST']:
            if not is_manager(request.user):
                return JsonResponse({'error': 'Недостаточно прав'}, status=403)
            return JsonResponse({'message': 'Для изменения используйте веб-форму /dashboard/grants/<id>/edit/'})
        if request.method == 'DELETE':
            if not is_admin(request.user):
                return JsonResponse({'error': 'Недостаточно прав'}, status=403)
            grant.delete()
            return JsonResponse({'message': 'Грант удален'})

    if request.method == 'GET':
        return JsonResponse({'results': [grant_to_dict(grant) for grant in filtered_grants(request)]})
    if request.method == 'POST':
        if not is_manager(request.user):
            return JsonResponse({'error': 'Недостаточно прав'}, status=403)
        return JsonResponse({'message': 'Создание доступно через форму /dashboard/grants/new/'})
    return JsonResponse({'error': 'Метод не поддерживается'}, status=405)


@login_required
def api_profile(request):
    if request.method == 'GET':
        user = request.user
        return JsonResponse({
            'id': user.id,
            'full_name': user.full_name,
            'email': user.email,
            'role': user.role,
            'university': user.university,
            'faculty': user.faculty,
            'course': user.course,
            'education_level': user.education_level,
            'interests': user.interests,
        })
    return JsonResponse({'message': 'Обновление профиля доступно через /profile/'})


@login_required
def api_my_applications(request):
    data = [{
        'id': item.id,
        'grant': item.grant.title,
        'status': item.get_status_display(),
        'submitted_at': item.submitted_at.isoformat(),
    } for item in request.user.applications.select_related('grant')]
    return JsonResponse({'results': data})


@login_required
def api_saved_grants(request):
    if request.method == 'GET':
        data = [grant_to_dict(item.grant) for item in SavedGrant.objects.filter(user=request.user).select_related('grant', 'grant__category')]
        return JsonResponse({'results': data})
    return JsonResponse({'message': 'Сохранение доступно через POST /saved-grants/<id>/'})


@csrf_exempt
@login_required
def api_create_application(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Метод не поддерживается'}, status=405)
    grant = get_object_or_404(Grant, pk=request.POST.get('grant_id'))
    form = ApplicationForm(request.POST, request.FILES)
    if not form.is_valid():
        return JsonResponse({'errors': form.errors}, status=400)
    try:
        application = form.save(commit=False)
        application.user = request.user
        application.grant = grant
        application.save()
    except IntegrityError:
        return JsonResponse({'error': 'Заявка на этот грант уже существует'}, status=400)
    return JsonResponse({'id': application.id, 'status': application.get_status_display()}, status=201)


@csrf_exempt
@login_required
def api_update_application_status(request, pk):
    if request.method not in ['PUT', 'POST']:
        return JsonResponse({'error': 'Метод не поддерживается'}, status=405)
    application = get_object_or_404(Application, pk=pk)
    if not is_manager(request.user):
        return JsonResponse({'error': 'Недостаточно прав'}, status=403)
    if request.user.role == User.Role.ORGANIZATION and application.grant.created_by != request.user and not request.user.is_staff:
        return JsonResponse({'error': 'Недостаточно прав'}, status=403)

    payload = request.POST
    if request.method == 'PUT':
        try:
            payload = json.loads(request.body.decode() or '{}')
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Некорректный JSON'}, status=400)
    status = payload.get('status')
    if status not in dict(Application.Status.choices):
        return JsonResponse({'error': 'Некорректный статус'}, status=400)
    application.status = status
    application.save(update_fields=['status'])
    Notification.objects.create(
        user=application.user,
        message=f'Статус заявки на "{application.grant.title}" изменен: {application.get_status_display()}.',
    )
    return JsonResponse({'id': application.id, 'status': application.get_status_display()})


@login_required
def api_recommendations(request):
    return JsonResponse({'results': [grant_to_dict(grant) for grant in recommended_grants_for(request.user)]})


@login_required
def api_notifications(request):
    return JsonResponse({'results': [{'message': n.message, 'is_read': n.is_read, 'created_at': n.created_at.isoformat()} for n in request.user.notifications.all()]})
